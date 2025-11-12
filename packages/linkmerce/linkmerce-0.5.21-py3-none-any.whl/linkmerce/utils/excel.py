from __future__ import annotations

from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from openpyxl import _ZipFileFileProtocol


def filter_warnings():
    import warnings
    warnings.filterwarnings("ignore", module="openpyxl.*")


def to_unique_headers(headers: list[str]) -> list[str]:
    unique = list()
    for header in headers:
        suffix = 1
        while header in unique:
            header = header + f'_{suffix}'
            suffix += 1
        unique.append(header)
    return unique


def csv2json(
        io: _ZipFileFileProtocol,
        header: int = 0,
        delimiter: str = ",",
        lineterminator: str = "\r\n",
        encoding: str | None = "utf-8",
    ) -> list[dict]:
    import os
    if isinstance(io, str) and os.path.exists(io):
        with open(io, 'r', encoding=encoding) as file:
            csv2json(file, header)

    import csv
    if isinstance(io, bytes):
        from io import BytesIO, TextIOWrapper
        io = TextIOWrapper(BytesIO(io), encoding=encoding)
    rows = list(csv.reader(io, delimiter=delimiter, lineterminator=lineterminator))
    header_row = to_unique_headers(rows[header])
    return [dict(zip(header_row, row)) for row in rows[(header+1):]]


def excel2json(
        io: _ZipFileFileProtocol,
        sheet_name: str | None = None,
        header: int = 1,
        warnings: bool = True
    ) -> list[dict]:
    from openpyxl import load_workbook
    from io import BytesIO
    if not warnings:
        filter_warnings()

    wb = load_workbook(BytesIO(io) if isinstance(io, bytes) else io)
    ws = wb.active if sheet_name is None else wb[sheet_name]

    headers = to_unique_headers([cell.value for cell in next(ws.iter_rows(min_row=header, max_row=header))])
    return [dict(zip(headers, row)) for row in ws.iter_rows(min_row=header+1, values_only=True)]
