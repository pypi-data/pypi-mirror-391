from typing import Dict, List, Union, Tuple
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from dataclasses import dataclass

@dataclass
class HeaderConfig:
    """Configuration for Excel headers including merge information"""
    text: str
    start_col: Union[int, str]
    end_col: Union[int, str] = None
    start_row: int = 1
    end_row: int = None
    alignment: dict = None
    style: dict = None

class ExcelWriter:
    """Class to handle writing complex Excel files with merged headers"""
    
    def __init__(self, filename: str):
        self.filename = filename
        self.workbook = openpyxl.Workbook()
        self.current_sheet = None
        
    def create_sheet(self, name: str) -> None:
        """Create a new worksheet"""
        if name in self.workbook.sheetnames:
            self.workbook.remove(self.workbook[name])
        self.current_sheet = self.workbook.create_sheet(name)
        
    def write_complex_headers(self, 
                            headers: List[HeaderConfig],
                            start_row: int = 1) -> None:
        """
        Write complex headers with merging
        
        Args:
            headers: List of HeaderConfig objects defining the headers
            start_row: Starting row for headers
        """
        if not self.current_sheet:
            raise ValueError("No active sheet. Call create_sheet() first.")
            
        for header in headers:
            # Convert column letters to numbers if needed
            start_col = header.start_col if isinstance(header.start_col, int) else \
                       openpyxl.utils.column_index_from_string(header.start_col)
            
            end_col = header.end_col if isinstance(header.end_col, int) else \
                     openpyxl.utils.column_index_from_string(header.end_col) if header.end_col else start_col
                     
            end_row = header.end_row if header.end_row else header.start_row
            
            # Write header text
            self.current_sheet.cell(header.start_row, start_col, header.text)
            
            # Merge cells if needed
            if start_col != end_col or header.start_row != end_row:
                merge_range = f"{get_column_letter(start_col)}{header.start_row}:"
                merge_range += f"{get_column_letter(end_col)}{end_row}"
                self.current_sheet.merge_cells(merge_range)
            
            # Apply styles if provided
            cell = self.current_sheet.cell(header.start_row, start_col)
            if header.alignment:
                cell.alignment = openpyxl.styles.Alignment(**header.alignment)
            if header.style:
                for style_attr, style_value in header.style.items():
                    setattr(cell, style_attr, style_value)
                    
    def write_dataframe(self,
                       df: pd.DataFrame,
                       start_row: int,
                       start_col: Union[int, str] = 1,
                       index: bool = False) -> None:
        """
        Write pandas DataFrame to Excel below headers
        
        Args:
            df: Pandas DataFrame to write
            start_row: Starting row for DataFrame
            start_col: Starting column (number or letter)
            index: Whether to write index
        """
        if isinstance(start_col, str):
            start_col = openpyxl.utils.column_index_from_string(start_col)
            
        # Write DataFrame
        for i, col in enumerate(df.columns):
            self.current_sheet.cell(start_row, start_col + i, str(col))
            
        for i, row in enumerate(df.values):
            for j, val in enumerate(row):
                self.current_sheet.cell(start_row + i + 1, start_col + j, val)
                
    def save(self) -> None:
        """Save the Excel file"""
        if "Sheet" in self.workbook.sheetnames:
            self.workbook.remove(self.workbook["Sheet"])
        self.workbook.save(self.filename)

if __name__ == "__main__":
    # Sample DataFrame
    df = pd.DataFrame({
        'Col1': [1, 2, 3],
        'Col2': [4, 5, 6],
        'Col3': [7, 8, 9],
        'Col4': [10, 11, 12]
    })
    
    # Create writer
    writer = ExcelWriter('complex_output.xlsx')
    writer.create_sheet('Sheet1')
    
    # Define complex headers
    headers = [
        HeaderConfig(
            text="Main Header 1",
            start_col="A",
            end_col="B",
            start_row=1,
            end_row=1,
            alignment={'horizontal': 'center'},
            style={'font': openpyxl.styles.Font(bold=True)}
        ),
        HeaderConfig(
            text="Main Header 2",
            start_col="C",
            end_col="D",
            start_row=1,
            end_row=1,
            alignment={'horizontal': 'center'},
            style={'font': openpyxl.styles.Font(bold=True)}
        ),
        HeaderConfig(
            text="Sub Header 1",
            start_col="A",
            start_row=2,
            alignment={'horizontal': 'center'}
        ),
        HeaderConfig(
            text="Sub Header 2",
            start_col="B",
            start_row=2,
            alignment={'horizontal': 'center'}
        )
    ]
    
    # Write headers and data
    writer.write_complex_headers(headers)
    writer.write_dataframe(df, start_row=3)
    writer.save()