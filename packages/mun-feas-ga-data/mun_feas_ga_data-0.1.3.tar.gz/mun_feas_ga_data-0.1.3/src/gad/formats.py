import abc
import collections
import csv
import openpyxl
import os.path
import re

from enum import IntEnum, StrEnum
from typing import Any, Iterable, NamedTuple, Optional


class DataError(Exception):
    """An error in GA data (not in code)"""


class Program(StrEnum):
    """An academic program in MUN Engineering."""

    CIVIL = "ENCV"
    COMPUTER = "ENCM"
    ELECTRICAL = "ENEL"
    MECHANICAL = "ENME"
    MECHATRONICS = "ENMT"
    ONAE = "ENON"
    PROCESSS = "ENPR"


class GA(IntEnum):
    """Graduate Attributes specified by the Canadian Engineering Accreditation Board.

    Every graduating Engineering student is supposed to possess these attributes.
    """

    KNOWLEDGE_BASE = 1
    PROBLEM_ANALYSIS = 2
    INVESTIGATION = 3
    DESIGN = 4
    TOOLS = 5
    TEAMWORK = 6
    COMMUNICATION = 7
    PROFESSIONALISM = 8
    IMPACTS = 9
    ETHICS_EQUITY = 10
    ECONOMICS = 11
    LIFELONG_LEARNING = 12


class Indicator(NamedTuple):
    """An indicator is a division of a Graduate Attribute.

    We don't measure GAs directly: they are too broad. Instead, we measure several
    indicators of achievement for each GA.
    """

    name: str
    description: str

    def __str__(self):
        return f"{self.name} ({self.description})"


class IndicatorLevel(StrEnum):
    """Indicators can be assessed at one of three levels of proficiency."""

    INTRODUCED = "I"
    DEVELOPED = "D"
    APPLIED = "A"

    @classmethod
    def parse(cls, s):
        try:
            return cls(s)
        except ValueError:
            pass

        try:
            return cls[s.upper()]
        except ValueError as ve:
            raise DataError(ve)


class Course(NamedTuple):
    """A course has a subject (e.g., "ENGI") and a code (e.g., "200W")."""

    # Course subject, e.g., "ECE"
    subject: str

    # Course code, e.g., "200W"
    code: str

    @staticmethod
    def parse(course: str) -> Optional["Course"]:
        if not course or "XXX" in course:
            return None

        subject, code = course.split()
        return Course(subject=subject, code=code)

    def __str__(self):
        return f"{self.subject} {self.code}"


class AssessmentTool(NamedTuple):
    """An assessment tool is an element of a course that can be used to provide insight
    into a student's achievement of an indicator.
    """

    # Course that contains the tool
    course: Course

    # Assessment method, e.g., "Final exam"
    method: str

    def __str__(self):
        return f"{self.course} {self.method}"


class ToolMapping(NamedTuple):
    """A mapping from one assessment tool to one indicator.

    Each assessment tool is used to provide insight into one student's achievement of
    one indicator at one proficiency level.
    """

    tool: AssessmentTool
    indicator: Indicator
    level: IndicatorLevel
    bins: list[int]

    def __str__(self):
        return f"{self.indicator.name}-{self.level}: {self.tool} ({self.bins})"


class CurriculumMap(NamedTuple):
    program: Program
    mappings: list[ToolMapping]

    # Course names contained in the map file
    course_names: dict[Course, str]

    # A mapping of GAs to indicators
    gas: dict[GA, set[Indicator]]

    # An index from indicators to assessment tool IDs (indices within `mappings`)
    indicator_index: dict[Indicator, int]

    def course_name(self, course):
        if course not in self.course_names:
            raise DataError(f"Course {course} not in curriculum map")
        else:
            return self.course_names[course]

    def mapping_for_tool(self, course: Course, tool_name: str) -> ToolMapping:
        for m in self.mappings:
            if m.tool.course == course and m.tool.method == tool_name:
                return m

        raise CurriculumMap.MissingMapping(tool_name, course, self)

    class MissingMapping(DataError):
        def __init__(self, tool_name, course, curriculum_map, *args, **kwargs):
            super().__init__(*args, **kwargs)

            self.tool_name = tool_name
            self.course = course
            self.curriculum_map = curriculum_map

        def __str__(self):
            course_mappings = ", ".join(
                [
                    m.tool.method
                    for m in self.curriculum_map.mappings
                    if m.tool.course == self.course
                ]
            )

            return f"No assessment tool '{self.tool_name}' found for {self.course} (options: {course_mappings})"


class FileFormat(abc.ABC):
    """A file format describes a way of representing assessment tool data within a file.

    Such a format should have an ability to check whether a filename could potentially
    describe AT data in this format and a method to parse that file.
    """

    @classmethod
    @abc.abstractmethod
    def filename_matches(cls, filename: str) -> bool:
        """Whether or not this file format makes sense for a filename.

        Parameters
        ----------
        filename : str
            The filename to be checked
        """

        return False

    @classmethod
    def open_excel_active_worksheet(cls, filename):
        """Open an Excel workbook's current active worksheet.

        The Excel workbooks we use mostly just have one worksheet, the active one.
        """

        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active

        if sheet is None:
            raise DataError(f"No active sheet in workbook {filename}")

        return sheet

    @classmethod
    @abc.abstractmethod
    def parse(cls, filename: str, curriculum_map: Optional[CurriculumMap]) -> Any:
        """Parse the contents of a file."""

        pass


class CurriculumMapFile(FileFormat):
    """A curriculum map file maps assessment tools to indicators.

    Each row contains the details of one assessment tool in one course being mapped to
    one indicator (or, in some cases, a non-assessed instance of teaching that relates
    to an indicator):

    - GA (number)
    - Graduate Attribute (name)
    - Indicator (code)
    - Indicator Description (string)
    - Level (Introduced, Developed or Applied)
    - Course (e.g., "ENGI 200W")
    - Course Title
    - Assessed ("Yes" or "No")
    - Method of Assessment (a.k.a., assessment tool name)
    - Bins (comma-separated list)
    - Term Collected (Fall, Winter or Spring)
    - Unit (probably same as subject?)
    - Co20xx (which we'll ignore)
    - Notes
    """

    filename_pattern = re.compile(r"(?P<program>EN[A-Z]{2}).*\.xlsx")

    @classmethod
    def filename_matches(cls, filename: str) -> bool:
        return cls.filename_pattern.match(filename) is not None

    @classmethod
    def parse(cls, filename: str, curriculum_map=None) -> CurriculumMap:
        assert curriculum_map is None  # we're supposed to be parsing the map!

        m = cls.filename_pattern.match(os.path.basename(filename))
        assert m is not None
        details = m.groupdict()
        program = Program(details["program"])

        #
        # Parse assessment tool mappings from the curriculum map's rows
        #
        sheet = cls.open_excel_active_worksheet(filename)
        headers = [str(c[0].value) for c in sheet.iter_cols(min_row=1, max_row=1)]

        course_names = {}
        mappings = []
        gas = collections.defaultdict(set)
        indicator_index = {}

        done = False
        for row in sheet.iter_rows(min_row=2):

            def field(name: str) -> str:
                if name not in headers:
                    raise DataError(
                        f"Curriculum map does not contain '{name}';"
                        + f" headers are: {', '.join(headers)}"
                    )

                value = row[headers.index(name)].value
                return str(value) if value is not None else ""

            # Ensure that empty rows only come at the end
            if row[0].value is None:
                done = True
                continue
            else:
                assert not done  # can't have empty rows before non-empty rows

            # Skip rows from the old curriculum maps that don't describe an assessment
            if "Assessed" in headers and field("Assessed") != "Yes":
                continue

            # Get the GA number (1–12) and indicator name
            ga = GA(int(field("GA")))
            indicator = Indicator(
                name=field("Indicator #"), description=field("Indicator Description")
            )

            # Check that the course (if present) is named consistently
            course = Course.parse(field("Course #"))
            if not course:
                # Skip rows with TBD courses (they haven't been mapped yet)
                continue

            course_name = field("Course Title")
            if course not in course_names:
                course_names[course] = course_name
            elif course_names[course] != course_name:
                raise DataError(
                    f"{course} has inconsistent names in {filename} ({course_name} vs {course_names[course]})"
                )

            # Get details of the assessment tool itself
            tool_name = field("Method of Assessment").capitalize()
            tool = AssessmentTool(course=course, method=tool_name)
            bins = [int(b) for b in field("Bins").split(",") if b]

            gas[ga].add(indicator)
            indicator_index[indicator] = len(mappings)
            mappings.append(
                ToolMapping(
                    tool=tool,
                    indicator=indicator,
                    level=IndicatorLevel.parse(field("Level")),
                    bins=bins,
                )
            )

        return CurriculumMap(
            program=program,
            mappings=mappings,
            course_names=course_names,
            gas=gas,
            indicator_index=indicator_index,
        )


class ATResults(NamedTuple):
    """A set of results for GA assessment tools.

    This named tuple contains a set of names for the assessment tool data it contains
    (e.g., "Final Exam" or "Design Project (Presentation Quality)"), and the results
    for those assessment tools organized as a list of results per student:


    (student ID if known, assessment tool name -> [0.0, 100.0] | None)
    """

    tools: set[str]
    results: list[tuple[int, dict[str, float | None]]]


class GADataFile(NamedTuple):
    """A file that contains GA assessment tool results.

    This file includes metadata taken from file as well as assessment tool results.
    """

    course: Course

    academic_year: int
    semester: int

    results: ATResults

    def __str__(self):
        course = str(self.course)
        semester_name = ["Fall", "Winter", "Spring"][self.semester - 1]
        year = f"{self.academic_year:04}–{(self.academic_year + 1) % 100:02}"

        return f"{course} ({semester_name} {year}): {len(self.results.results)} results"


class GADataFileFormat(FileFormat):
    @classmethod
    @abc.abstractmethod
    def parse(
        cls, filename: str, curriculum_map: Optional[CurriculumMap]
    ) -> GADataFile:
        """Parse a file containing GA data."""

        pass


class ATsheet(GADataFileFormat):
    """The ATsheet format contains assessment tool data as a first-class primitive.

    A file in this format is named something like ECE1234-F2025-GAdata.csv and contains
    one row per student, each of which contains a student ID and that student's result
    for each assessment tool, e.g.:

    ```csv
    Student number,Final Exam, Assignments
    202412345,89,95
    202523456,75,84
    ```
    """

    filename_pattern = re.compile(
        r"((?P<subject>[A-Z]+)(?P<course_code>[0-9]+)_(?P<semester>[A-Z])(?P<year>[0-9]{4})_GAdata)\.(?P<format>(csv)|(xlsx))"
    )

    student_id_names = [
        "OrgDefinedId",
        "Student number",
        "Student Number",
    ]

    @classmethod
    def filename_matches(cls, filename: str) -> bool:
        return cls.filename_pattern.match(filename) is not None

    @classmethod
    def parse(
        cls, filename: str, curriculum_map: Optional[CurriculumMap]
    ) -> GADataFile:
        del curriculum_map

        m = cls.filename_pattern.match(os.path.basename(filename))
        assert m is not None
        details = m.groupdict()

        course = Course(subject=details["subject"], code=details["course_code"])
        year = int(details["year"])
        semester = {"F": 1, "W": 2, "S": 3}[details["semester"]]

        match details["format"]:
            case "csv":
                results = cls.parse_csv(filename)

            case "xlsx":
                results = cls.parse_excel(filename)

            case _:
                raise DataError(f"Invalid file format: {details['format']}")

        return GADataFile(
            course=course,
            semester=semester,
            academic_year=year if semester == 1 else year - 1,
            results=results,
        )

    @classmethod
    def parse_csv(cls, filename: str) -> ATResults:
        with open(filename, "r", encoding="utf-8-sig") as f:
            reader = csv.reader(f)

            first_row = next(reader)
            sid_name = first_row[0]
            if sid_name not in cls.student_id_names:
                raise DataError(f"Invalid student ID column name: '{sid_name}'")

            tool_names = [t.capitalize() for t in first_row[1:] if t]
            results = []

            for row in reader:
                sid = int(row[0])
                results.append((sid, dict(zip(tool_names, row[1:]))))

            return ATResults(tools=set(tool_names), results=results)

    @classmethod
    def parse_excel(cls, filename: str) -> ATResults:
        sheet = cls.open_excel_active_worksheet(filename)

        sid_name = sheet["A1"].value
        if sid_name not in cls.student_id_names:
            raise DataError(f"Invalid student ID column name: '{sid_name}'")

        tool_names = [
            str(col[0].value).capitalize()
            for col in sheet.iter_cols(2, sheet.max_column)
            if col[0].value
        ]
        results = []

        for row in range(2, sheet.max_row):
            sid = sheet.cell(row=row, column=1).value
            if isinstance(sid, int):
                pass
            elif isinstance(sid, str):
                sid = int(sid)
            else:
                raise DataError(f"Invalid student ID: {sid}")

            tool_values = [
                cls.cell_value(sheet.cell(row=row, column=col))
                for col in range(2, sheet.max_column)
            ]

            results.append((sid, dict(zip(tool_names, tool_values))))

        return ATResults(tools=set(tool_names), results=results)

    @classmethod
    def cell_value(cls, cell) -> float | None:
        val = cell.value

        if isinstance(val, str):
            if val.endswith("%"):
                val = val[:-1]

            if val:
                return float(val)
            else:
                return None
        elif val is None:
            return val
        else:
            return float(val)


class FEAMS(GADataFileFormat):
    """The FEAMS format contains one assessment tool's data per file.

        A file in this format is named something like 2025-SPRING-TOOLS1-D.xlsx and contains
        one assessment tool's data in cells B14–Bxxx, with some metadata in cells B1–B13
        (and labels in A1–A14), e.g.:

        ```csv
        Department undergoing Accreditation,CE
        Course Faculty/Department,ECE
        Course No.,4500
        Course Title,Microprocessors
        Semester,Spring
        Term,4
        Course Instructor
        Assessment Tool,Labs
        Bin Level for "No Grade Available",0
        Bin Level for "Below Expectations",70
        Bin Level for "Marginally Meets Expectations",80
        Bin Level for "Meets Expectations",90
        Bin Level for "Exceeds Expectations",100
        Assessment Tool Data (Enter -1 for "Blank" or "No Grade Available" Entries,71
        ,81
        ,-1
        ,76
        ,82
        ,94
    ```
    """

    filename_pattern = re.compile(
        r"((?P<year>[0-9]{4})-(?P<semester>[A-Z]+)-(?P<indicator>[A-Z0-9]+)-(?P<level>[IDA]))\.xlsx"
    )

    semesters = ["FALL", "WINTER", "SPRING"]

    @classmethod
    def filename_matches(cls, filename: str) -> bool:
        return cls.filename_pattern.match(filename) is not None

    @classmethod
    def parse(
        cls, filename: str, curriculum_map: Optional[CurriculumMap]
    ) -> GADataFile:
        if not curriculum_map:
            raise RuntimeError(
                f"Cannot parse FEAMS data file ({filename}) without a curriculum map"
            )

        m = cls.filename_pattern.match(os.path.basename(filename))
        assert m is not None
        details = m.groupdict()

        semester = details["semester"]
        if semester not in cls.semesters:
            raise DataError(f"Invalid semester name: '{semester}'")

        semester = cls.semesters.index(semester) + 1
        academic_year = int(details["year"]) - int(semester > 1)

        sheet = cls.open_excel_active_worksheet(filename)

        course = Course(subject=str(sheet["B2"].value), code=str(sheet["B3"].value))
        tool_name = str(sheet["B8"].value).capitalize()

        # Check that this data corresponds to an entry in the curriculum map
        mapping = curriculum_map.mapping_for_tool(course, tool_name)

        # Confirm that the expected bins match the curriculum map
        bins = [cls.int_value(sheet.cell(row=i + 10, column=2)) for i in range(4)]

        if mapping.bins and bins != mapping.bins:
            raise DataError(
                f"Bins in {filename} ({bins}) do not match those in curriculum map ({mapping.bins})"
            )

        # Confirm the indicator level also matches
        level = IndicatorLevel(details["level"])
        if level != mapping.level:
            raise DataError(
                f"Assessment level in {filename} ({level}) does not match curriculum map ({mapping})"
            )

        # Parse the actual data results
        results = []

        for row in sheet.iter_rows(min_row=14, max_row=sheet.max_row):
            results.append((mapping.indicator, {tool_name: cls.int_value(row[1])}))

        return GADataFile(
            course=course,
            academic_year=academic_year,
            semester=semester,
            results=ATResults(tools=set([tool_name]), results=results),
        )

    @classmethod
    def dump_files(
        cls, output_dir: str, at_data: GADataFile, curriculum_map: CurriculumMap
    ):
        course = at_data.course

        long_semester_name = cls.semesters[at_data.semester - 1]
        filename_prefix = f"{at_data.academic_year}-{long_semester_name}"

        for tool in at_data.results.tools:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            assert sheet is not None

            if course not in curriculum_map.course_names:
                print(f"{course} not in curriculum map; skipping {tool}")
                continue

            try:
                mapping = curriculum_map.mapping_for_tool(course, tool)
            except CurriculumMap.MissingMapping:
                # Skip tools that don't appear in the curriculum map
                print(f"{course} {tool} not in curriculum map; skipping")
                continue

            assert curriculum_map.program
            sheet.append(["Academic program", curriculum_map.program])
            sheet.append(["Course subject", course.subject])
            sheet.append(["Course code", course.code])
            sheet.append(["Course title", curriculum_map.course_names[course]])
            sheet.append(["Semester", long_semester_name.title()])
            sheet.append(["Term", at_data.course.code[0]])  # TODO: fix?
            sheet.append(["Course instructor", "???"])  # TODO: fix?
            sheet.append(["Assessment Tool", tool])

            # TODO: get bin levels from... curriculum map?
            bin_levels = [0, 70, 80, 90, 100]
            for j, bin_name in enumerate(
                [
                    "No Grade Available",
                    "Below Expectations",
                    "Marginally Meets Expectations",
                    "Meets Expectations",
                    "Exceeds Expectations",
                ]
            ):
                sheet.append([f'Bin Level for "{bin_name}"', bin_levels[j]])

            for _, results in at_data.results.results:
                grade = results.get(tool, None)
                sheet.append(["", -1 if grade is None else grade])

            # TODO: import curriculum map to get real indicator
            indicator = f"{mapping.indicator.name}-{mapping.level}"
            filename = os.path.join(output_dir, f"{filename_prefix}-{indicator}.xlsx")

            os.makedirs(output_dir, exist_ok=True)
            workbook.save(filename)

    @classmethod
    def int_value(cls, cell) -> int | None:
        val = cell.value

        if val == -1:
            return None
        elif isinstance(val, str):
            if val:
                return int(val)
        elif val is None:
            return val
        else:
            return int(val)


all_formats: list[type[FileFormat]] = [
    ATsheet,
    FEAMS,
]


def parse(
    files_or_directories: Iterable[str], curriculum_map: Optional[CurriculumMap]
) -> list[GADataFile]:
    """Parse all of the files contained within a set of files and/or directories.

    Parameters
    ----------
    files_or_directories
        A collection of files and/or directories.

    Returns
    -------
    Parsed assessment tool results from each file contained within the files and/or
    directories passed in (via recursive search).
    """

    all_files = []

    for filename in collect_files(files_or_directories):
        assert os.path.isfile(filename)

        for file_format in all_formats:
            if file_format.filename_matches(os.path.basename(filename)):
                all_files.append(file_format.parse(filename, curriculum_map))

    return all_files


def collect_files(files_or_directories: Iterable[str]) -> list[str]:
    """Collect all of the files from a set of files and/or directories."""

    all_files = []

    for fd in files_or_directories:
        if not os.path.exists(fd):
            raise DataError(f"{fd} does not exist")

        if os.path.isfile(fd):
            all_files.append(fd)
        elif os.path.isdir(fd):
            for dirpath, _, filenames in os.walk(fd):
                all_files += [os.path.join(dirpath, name) for name in filenames]
        else:
            raise DataError(f"{fd} is neither a file nor a directory")

    return all_files
