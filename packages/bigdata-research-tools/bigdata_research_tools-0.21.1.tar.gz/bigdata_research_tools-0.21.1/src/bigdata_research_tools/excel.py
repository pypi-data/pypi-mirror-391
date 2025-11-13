"""
Module for managing Excel workbook operations.

Copyright (C) 2024, RavenPack | Bigdata.com. All rights reserved.
"""

from logging import Logger, getLogger
from typing import List, Tuple

import pandas as pd

from bigdata_research_tools.settings import (
    check_libraries_installed,
    get_resources_path,
)

logger: Logger = getLogger(__name__)

# Constants for Excel formatting
MIN_COLUMN_WIDTH = 12
MAX_COLUMN_WIDTH = 75
ROW_OFFSET = 3
COLUMN_OFFSET = 1
LOGO_SCALE_FACTOR = 0.2
LOGO_ROW_HEIGHT = 30
LOGO_COLUMN_WIDTH = 5


def check_excel_dependencies() -> bool:
    """
    Check if the required Excel dependencies are installed.
    Will look for the `openpyxl` and `pillow` packages.
    """
    return check_libraries_installed(["openpyxl", "PIL"])


class ExcelManager:
    """Class for managing Excel workbook operations."""

    def __init__(
        self,
        min_column_width: int = MIN_COLUMN_WIDTH,
        max_column_width: int = MAX_COLUMN_WIDTH,
        row_offset: int = ROW_OFFSET,
        column_offset: int = COLUMN_OFFSET,
    ):
        """Initialize Excel manager with formatting parameters."""
        from openpyxl.styles import Border, Side

        self.min_column_width = min_column_width
        self.max_column_width = max_column_width
        self.row_offset = row_offset
        self.column_offset = column_offset

        self.logo_path = f"{get_resources_path()}/bigdata-by-ravenpack-logo.png"

        self.thick_border = Border(left=Side(style="thick"), right=Side(style="thick"))

    def save_workbook(
        self,
        df_args: List[Tuple[pd.DataFrame, str, Tuple[int, int]]],
        workbook_path: str,
    ) -> None:
        """Save DataFrames to Excel workbook."""
        logger.info(f"Saving workbook to `{workbook_path}`")
        with pd.ExcelWriter(workbook_path) as excel_writer:
            for df, sheet_name, freeze_panes in df_args:
                df.to_excel(
                    excel_writer,
                    sheet_name=sheet_name,
                    index=False,
                    startrow=self.row_offset - 2,
                    startcol=self.column_offset,
                    freeze_panes=freeze_panes,
                )
        self.beautify_workbook(workbook_path)

    def beautify_workbook(self, workbook_path: str) -> None:
        """Apply formatting to Excel workbook."""
        from openpyxl import load_workbook

        wb = load_workbook(workbook_path)
        for sheet_name in wb.sheetnames:
            self._beautify_worksheet(wb[sheet_name])
        wb.save(workbook_path)

    def _add_branding(self, sheet) -> None:
        """Add branding elements to worksheet with LibreOffice compatibility."""
        from openpyxl.drawing.image import Image

        img = Image(self.logo_path)

        img.width = int(img.width * LOGO_SCALE_FACTOR)
        img.height = int(img.height * LOGO_SCALE_FACTOR)

        img.anchor = "A1"

        sheet.add_image(img)

        sheet.row_dimensions[1].height = LOGO_ROW_HEIGHT
        sheet.column_dimensions["A"].width = LOGO_COLUMN_WIDTH

    def _adjust_column_widths(self, sheet) -> None:
        """Adjust column widths based on content."""
        for column_cells in sheet.columns:
            max_length = 0
            column = column_cells[0].column_letter
            for cell in column_cells[self.column_offset + 1 :]:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(
                max(max_length, self.min_column_width), self.max_column_width
            )
            sheet.column_dimensions[column].width = adjusted_width + 2

    def _apply_zebra_striping(self, sheet) -> None:
        """Apply zebra striping to worksheet."""
        from openpyxl.styles import PatternFill

        stripe_fill = PatternFill(
            start_color="DDDDDD", end_color="DDDDDD", fill_type="solid"
        )
        for idx, row in enumerate(
            sheet.iter_rows(min_row=self.row_offset, min_col=self.column_offset + 1)
        ):
            if idx % 2 == 1:
                for cell in row:
                    cell.fill = stripe_fill

    def _beautify_worksheet(self, sheet) -> None:
        """Apply formatting to worksheet."""
        sheet.sheet_view.showGridLines = False

        # Format cells
        for row in sheet.iter_rows(min_row=self.row_offset, min_col=self.column_offset):
            for cell in row:
                self._format_cell(cell)

        # Add branding and format header
        self._add_branding(sheet)
        self._format_header(sheet)

        # Format columns and apply striping
        self._adjust_column_widths(sheet)
        self._apply_zebra_striping(sheet)
        self._format_special_columns(sheet)

    def _format_cell(self, cell) -> None:
        """Format individual cell."""
        from openpyxl.styles import Alignment, Font

        cell.font = Font(size=12)
        cell_length = len(str(cell.value)) if cell.value else 0
        h_alignment = "center" if cell_length < self.max_column_width else "left"
        cell.alignment = Alignment(horizontal=h_alignment, vertical="center")
        if cell.value == 0:
            cell.value = "-"

    def _format_header(self, sheet) -> None:
        """Format header row."""
        from openpyxl.styles import Alignment, Font, PatternFill

        header_row = sheet[self.row_offset - 1][self.column_offset :]
        for cell in header_row:
            cell.font = Font(size=12, bold=True, color="000000")
            cell.fill = PatternFill(
                start_color="8CB3F8", end_color="8CB3F8", fill_type="solid"
            )
            cell.alignment = Alignment(
                wrap_text=True, horizontal="center", vertical="top"
            )
        header_ref = f"{header_row[0].coordinate}:{header_row[-1].coordinate}"
        sheet.auto_filter.ref = header_ref

    def _format_special_columns(self, sheet) -> None:
        """Format first and last columns."""
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter

        # Format first column
        for cell in sheet[self.column_offset + 1]:
            cell.font = Font(size=12, bold=True)

        # Format last column
        last_col = sheet.max_column
        last_col_letter = get_column_letter(last_col)
        for cell in sheet[last_col_letter][self.row_offset - 2 :]:
            cell.font = Font(size=12, bold=True)
            cell.border = self.thick_border
