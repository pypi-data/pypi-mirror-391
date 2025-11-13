from abc import ABC, abstractmethod
from io import BytesIO
from typing import List, Tuple, Optional

import xlrd
from openpyxl.reader.excel import load_workbook
from pydantic import Field, BaseModel

from doc_parser.dom_parser.domtree.domtree import DomTree, Node, DomTreeModel
from doc_parser.dom_parser.parsers.base import BaseConverter
from doc_parser.dom_parser.parsers.pdf.extend.table.TableBlockExtend import TableBlockExtend
from doc_parser.dom_parser.parsers.pdf.extend.text.TextBlockExtend import TextBlockExtend
from doc_parser.dom_parser.parsers.pdf.layout.Blocks import Blocks
from doc_parser.dom_parser.parsers.pdf.table.Cell import Cell
from doc_parser.dom_parser.parsers.pdf.table.Row import Row
from doc_parser.dom_parser.parsers.pdf.table.TableBlock import TableBlock
from doc_parser.dom_parser.parsers.pdf.text.TextBlock import TextBlock


class CellData(BaseModel):
    """表示 Excel 单元格的数据结构

    用于存储 Excel 单元格的相关信息，包括单元格值、行列索引和边界框。

    Attributes:
        value (str): 单元格值，表示单元格中的内容
        row_index (int): 行索引，表示单元格所在的行
        col_index (int): 列索引，表示单元格所在的列
        bbox (List[int]): 单元格边界框，表示单元格在页面上的位置和大小，格式为 [x1, y1, x2, y2]
        is_merged_cell (bool): 是否是合并单元格
        merged_width (int): 合并单元格的宽度（列数）
        merged_height (int): 合并单元格的高度（行数）
    """
    value: str  # 单元格值
    row_index: int  # 行索引，继承自row
    col_index: int  # 列索引
    bbox: List[int]  # 单元格边界框 [x1, y1, x2, y2]
    is_merged_cell: bool = False  # 是否是合并单元格
    merged_width: Optional[int]   # 合并单元格的宽度（列数）
    merged_height: Optional[int]  # 合并单元格的高度（行数）


class RowData(BaseModel):
    """表示 Excel 行的数据结构

    用于存储 Excel 行的相关信息，包括单元格列表和行索引。
    每行包含多个单元格，每个单元格由 CellData 对象表示。

    Attributes:
        cells (List[CellData]): 单元格列表，存储行中的所有单元格
        row_index (int): 行索引，表示该行在工作表中的位置
    """
    cells: List[CellData] = Field(default_factory=list)  # 单元格列表
    row_index: int = 0  # 行索引


class SheetData(BaseModel):
    """表示 Excel 工作表的数据结构

    用于存储 Excel 工作表的相关信息，包括工作表名称、行列表和页码。
    每个工作表包含多行数据，每行数据由 RowData 对象表示。

    Attributes:
        name (str): 工作表名称
        rows (List[RowData]): 行列表，存储工作表中的所有行
        page (int): 工作表所在的页码
    """
    name: str  # 工作表名称
    rows: List[RowData] = Field(default_factory=list)  # 行列表
    page: int  # 页码


class ExcelData(BaseModel):
    """表示 Excel 文件的数据结构

    用于存储整个 Excel 文件的数据，包括所有工作表。
    每个 Excel 文件包含多个工作表，每个工作表由 SheetData 对象表示。

    Attributes:
        sheets (List[SheetData]): 工作表列表，存储 Excel 文件中的所有工作表
    """
    sheets: Optional[List[SheetData]] = Field(default_factory=list)  # 工作表列表


class ExcelBaseConverter(BaseConverter, ABC):
    """Excel 转换器"""

    def __init__(self, stream: bytes):
        self.stream = stream

    def dom_tree_parse(self, start: int = 0, end: int = None, pages: list = None, **kwargs):
        # 解析 Excel 文件
        excel_data = self._parse_excel(self.stream)

        dom_tree = DomTree()

        # 遍历每个工作表，为每个工作表创建一个表格节点
        for i,sheet in enumerate(excel_data.sheets):
            table_block = self._build_table_block(sheet)
            table_block_extend = TableBlockExtend(table_block)
            table_block_extend.page_num = [i]
            # 将表格添加到 DOM 树中
            node = Node(table_block_extend, None, None)
            node.order_num_str = "1"
            dom_tree.root.add_child(node)

        return DomTreeModel(dom_tree = dom_tree)

    @abstractmethod
    def _parse_excel(self, byte_data) -> ExcelData:
        """解析 Excel 文件，返回 ExcelData 对象"""
        raise NotImplementedError("子类必须实现解析 Excel 文件的方法")

    def _build_table_block(self, sheet: SheetData):
        """构建 TableBlock 对象"""
        # 创建 TableBlock
        table_block = TableBlock()

        # 初始化表格边界框的最小和最大值
        min_x = float('inf')
        min_y = float('inf')
        max_x = float('-inf')
        max_y = float('-inf')

        # 遍历每一行
        for row_data in sheet.rows:
            # 创建行对象
            row = Row()

            # 如果行中有单元格，使用第一个和最后一个单元格的边界框来设置行的边界框
            if row_data.cells:
                first_cell = row_data.cells[0]
                last_cell = row_data.cells[-1]
                row_bbox = [
                    first_cell.bbox[0],
                    first_cell.bbox[1],
                    last_cell.bbox[2],
                    last_cell.bbox[3]
                ]
                row.update_bbox(row_bbox)
                row.height = row_bbox[3] - row_bbox[1]  # 行高

                # 更新表格边界框的最小和最大值
                min_x = min(min_x, row_bbox[0])
                min_y = min(min_y, row_bbox[1])
                max_x = max(max_x, row_bbox[2])
                max_y = max(max_y, row_bbox[3])

            # 遍历每个单元格
            for cell_data in row_data.cells:
                # 创建单元格对象
                raw_lines = [{'spans': [{'text': cell_data.value, 'bbox': cell_data.bbox}], 'bbox': cell_data.bbox}]
                raw_block = {'lines': raw_lines, 'bbox': cell_data.bbox}
                cell = Cell(raw=raw_block)
                cell.update_bbox(cell_data.bbox)
                cell.blocks = Blocks(instances=[TextBlockExtend(TextBlock(raw=raw_block))])  # 初始化单元格的块列表

                # # 处理合并单元格
                if cell_data.is_merged_cell:
                    cell.start_row = cell_data.row_index
                    cell.start_col = cell_data.col_index
                    cell.merged_cells = (cell_data.merged_height, cell_data.merged_width)
                else:
                    cell.start_row = cell_data.row_index
                    cell.start_col = cell_data.col_index

                # 将单元格添加到行中
                row.append(cell)

                # 更新表格边界框的最小和最大值
                min_x = min(min_x, cell_data.bbox[0])
                min_y = min(min_y, cell_data.bbox[1])
                max_x = max(max_x, cell_data.bbox[2])
                max_y = max(max_y, cell_data.bbox[3])

            # 将行添加到表格中
            table_block.append(row)

        # 设置表格的边界框（基于所有单元格的实际大小）
        if min_x != float('inf') and min_y != float('inf') and max_x != float('-inf') and max_y != float('-inf'):
            table_block.update_bbox([min_x, min_y, max_x, max_y])
        else:
            # 如果没有单元格，使用默认值
            table_block.update_bbox([0, 0, 0, 0])

        return table_block


class XlsExcelConverter(ExcelBaseConverter):

    def _parse_excel(self, byte_data) -> ExcelData:
        """解析 Excel 文件，返回 ExcelData 对象"""
        # 将字节数据转换为BytesIO对象
        byte_stream = BytesIO(byte_data)

        # 打开XLS文件
        workbook = xlrd.open_workbook(file_contents=byte_stream.read(), formatting_info=True)

        # 创建 ExcelData 对象
        excel_data = ExcelData(sheets=[])

        # 遍历每个工作表
        for sheet_index in range(workbook.nsheets):
            worksheet = workbook.sheet_by_index(sheet_index)
            sheet_name = worksheet.name

            # 获取合并单元格信息
            merged_cells = worksheet.merged_cells
            merged_cell_map = {}  # 用于存储每个单元格是否是合并单元格的一部分

            # 处理合并单元格信息
            for (row_start, row_end, col_start, col_end) in merged_cells:
                # 对于每个合并单元格，记录主单元格（左上角单元格）和被合并的单元格
                main_cell = (row_start, col_start)
                for row in range(row_start, row_end):
                    for col in range(col_start, col_end):
                        if (row, col) != main_cell:
                            # 被合并的单元格指向主单元格
                            merged_cell_map[(row, col)] = main_cell

            # 创建工作表的行列表
            rows = []

            # 遍历每一行
            for row_index in range(worksheet.nrows):
                # 创建行的单元格列表
                cells = []

                # 遍历每一列
                for col_index in range(worksheet.ncols):
                    # 检查当前单元格是否是被合并的单元格
                    if (row_index, col_index) in merged_cell_map:
                        # 如果是被合并的单元格，跳过
                        continue

                    # 获取单元格值
                    cell_value = worksheet.cell_value(row_index, col_index)
                    cell_type = worksheet.cell_type(row_index, col_index)

                    # 根据单元格类型处理值
                    if cell_type == xlrd.XL_CELL_DATE:
                        # 处理日期类型
                        date_tuple = xlrd.xldate_as_tuple(cell_value, workbook.datemode)
                        if date_tuple[3] == 0 and date_tuple[4] == 0 and date_tuple[5] == 0:
                            # 只有日期，没有时间
                            cell_value = f"{date_tuple[0]}-{date_tuple[1]:02d}-{date_tuple[2]:02d}"
                        else:
                            # 日期和时间
                            cell_value = f"{date_tuple[0]}-{date_tuple[1]:02d}-{date_tuple[2]:02d} {date_tuple[3]:02d}:{date_tuple[4]:02d}:{date_tuple[5]:02d}"

                    # 检查当前单元格是否是合并单元格的主单元格
                    is_merged_main_cell = False
                    merged_width = 1
                    merged_height = 1

                    for (row_start, row_end, col_start, col_end) in merged_cells:
                        if row_index == row_start and col_index == col_start:
                            is_merged_main_cell = True
                            merged_width = col_end - col_start
                            merged_height = row_end - row_start
                            break

                    # 创建单元格边界框（考虑合并单元格）
                    cell_width = 10  # 单元格宽度
                    cell_height = 10  # 单元格高度
                    if is_merged_main_cell:
                        bbox = [
                            col_index * cell_width,
                            row_index * cell_height,
                            (col_index + merged_width) * cell_width,
                            (row_index + merged_height) * cell_height
                        ]
                    else:
                        bbox = [
                            col_index * cell_width,
                            row_index * cell_height,
                            (col_index + 1) * cell_width,
                            (row_index + 1) * cell_height
                        ]

                    # 创建 CellData 对象并添加到单元格列表
                    cells.append(CellData(
                        value=str(cell_value),
                        row_index=row_index,
                        col_index=col_index,
                        bbox=bbox,
                        is_merged_cell=is_merged_main_cell,
                        merged_width=merged_width,
                        merged_height=merged_height
                    ))

                # 创建 RowData 对象并添加到行列表
                rows.append(RowData(cells=cells, row_index=row_index))

            # 创建 SheetData 对象并添加到工作表列表
            excel_data.sheets.append(SheetData(name=sheet_name, rows=rows, page=sheet_index))

        return excel_data

class XlsxExcelConverter(ExcelBaseConverter):

    def _parse_excel(self, byte_data) -> ExcelData:
        """解析 Excel 文件，返回 ExcelData 对象"""
        # 将字节数据转换为BytesIO对象
        byte_stream = BytesIO(byte_data)

        # 打开XLSX文件
        workbook = load_workbook(filename=byte_stream)

        # 创建 ExcelData 对象
        excel_data = ExcelData(sheets=[])

        # 遍历每个工作表
        for sheet_index, worksheet in enumerate(workbook.worksheets):
            sheet_name = worksheet.title

            # 获取合并单元格信息
            merged_cells = worksheet.merged_cells.ranges
            merged_cell_map = {}  # 用于存储每个单元格是否是合并单元格的一部分

            # 处理合并单元格信息
            for merged_range in merged_cells:
                # 获取合并单元格的范围
                min_row, min_col = merged_range.min_row - 1, merged_range.min_col - 1  # 转为0-based索引
                max_row, max_col = merged_range.max_row - 1, merged_range.max_col - 1  # 转为0-based索引

                # 对于每个合并单元格，记录主单元格（左上角单元格）和被合并的单元格
                main_cell = (min_row, min_col)
                for row in range(min_row, max_row + 1):
                    for col in range(min_col, max_col + 1):
                        if (row, col) != main_cell:
                            # 被合并的单元格指向主单元格
                            merged_cell_map[(row, col)] = main_cell

            # 创建工作表的行列表
            rows = []

            # 遍历每一行
            for row_index, row in enumerate(worksheet.rows):
                # 创建行的单元格列表
                cells = []

                # 遍历每一列
                for col_index, cell in enumerate(row):
                    # 检查当前单元格是否是被合并的单元格
                    if (row_index, col_index) in merged_cell_map:
                        # 如果是被合并的单元格，跳过
                        continue

                    # 获取单元格值
                    cell_value = cell.value

                    # 处理单元格值为 None 的情况
                    if cell_value is None:
                        cell_value = ""

                    # 处理日期类型
                    from datetime import datetime
                    if isinstance(cell_value, datetime):
                        if cell_value.hour == 0 and cell_value.minute == 0 and cell_value.second == 0:
                            # 只有日期，没有时间
                            cell_value = f"{cell_value.year}-{cell_value.month:02d}-{cell_value.day:02d}"
                        else:
                            # 日期和时间
                            cell_value = f"{cell_value.year}-{cell_value.month:02d}-{cell_value.day:02d} {cell_value.hour:02d}:{cell_value.minute:02d}:{cell_value.second:02d}"

                    # 检查当前单元格是否是合并单元格的主单元格
                    is_merged_main_cell = False
                    merged_width = 1
                    merged_height = 1

                    for merged_range in merged_cells:
                        min_row, min_col = merged_range.min_row - 1, merged_range.min_col - 1  # 转为0-based索引
                        max_row, max_col = merged_range.max_row - 1, merged_range.max_col - 1  # 转为0-based索引

                        if row_index == min_row and col_index == min_col:
                            is_merged_main_cell = True
                            merged_width = max_col - min_col + 1
                            merged_height = max_row - min_row + 1
                            break

                    # 创建单元格边界框（考虑合并单元格）
                    cell_width = 10  # 单元格宽度
                    cell_height = 10  # 单元格高度
                    if is_merged_main_cell:
                        bbox = [
                            col_index * cell_width,
                            row_index * cell_height,
                            (col_index + merged_width) * cell_width,
                            (row_index + merged_height) * cell_height
                        ]
                    else:
                        bbox = [
                            col_index * cell_width,
                            row_index * cell_height,
                            (col_index + 1) * cell_width,
                            (row_index + 1) * cell_height
                        ]

                    # 创建 CellData 对象并添加到单元格列表
                    cells.append(CellData(
                        value=str(cell_value),
                        row_index=row_index,
                        col_index=col_index,
                        bbox=bbox,
                        is_merged_cell=is_merged_main_cell,
                        merged_width=merged_width,
                        merged_height=merged_height
                    ))

                # 创建 RowData 对象并添加到行列表
                rows.append(RowData(cells=cells, row_index=row_index))

            # 创建 SheetData 对象并添加到工作表列表
            excel_data.sheets.append(SheetData(name=sheet_name, rows=rows, page=sheet_index))

        return excel_data