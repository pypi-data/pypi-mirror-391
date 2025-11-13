# -*- coding: utf-8 -*-
# ===============================================================
#
#    Copyright (C) 2024 Beike, Inc. All Rights Reserved.
#
#    @Create Author : luxu
#    @Create Time   : 2024/11/7
#    @Description   :
#
# ===============================================================
from io import BytesIO

import openpyxl

from doc_parser.layout_parser.layout.simple_block import SimpleBlock
from services.constants import TEXT
from services.layout_parse_utils import get_s3_links_for_simple_block_batch


def layout_parse(byte_data):
    # 将字节数据转换为BytesIO对象
    byte_stream = BytesIO(byte_data)

    # 打开XLSX文件
    workbook = openpyxl.load_workbook(byte_stream)

    # 用于存储所有内容的字符串
    all_content = ""

    # 遍历每个工作表
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        all_content += f"Sheet: {sheet}\n"

        # 遍历每一行
        for row in worksheet.iter_rows():
            # 遍历每一列
            for cell in row:
                # 将单元格的值拼接到字符串中
                all_content += str(cell.value) + "\t"
            all_content += "\n"
        all_content += "\n"

    result_text = all_content
    result_json = get_s3_links_for_simple_block_batch([SimpleBlock(type=TEXT, text=result_text)])
    return result_json, result_text
