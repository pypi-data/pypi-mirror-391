#  coding=utf-8
#
#  Author: Ernesto Arredondo Martinez (ernestone@gmail.com)
#  Created: 31/03/2019
#  Copyright (c)
import io

from .utils_logging import get_file_logger
from openpyxl import load_workbook


def xls_sheet(file_xls, sheet_name=None, logger=None, **workbook_args):
    """
    Devuelve una hoja del fichero excel especificado.
    Si no se especifica nombre de hora (sheet_name) la primera que encuentre.
    Se podran especificar argumentos extras para la función load_workbook() pero
    si no, por defecto en modo read_only=True y data_only=True

    Args:
        logger:
        file_xls (str): path fichero excel
        sheet_name (str=None): nombre hoja a devolver. Por defecto la primera que encuentre
        **workbook_args: Extra arguments para funcion load_workbook()

    Returns:
        openpyxl.worksheet
    """
    if not logger:
        logger = get_file_logger()

    read_only = workbook_args.pop('read_only', True)

    if read_only:
        with open(file_xls, "rb") as f:
            in_mem_file = io.BytesIO(f.read())
        a_wb = load_workbook(in_mem_file,
                             read_only=read_only,
                             data_only=workbook_args.pop('data_only', True))
    else:
        a_wb = load_workbook(file_xls,
                             read_only=read_only,
                             data_only=workbook_args.pop('data_only', True))

    sh1 = a_wb.sheetnames[0]
    if sheet_name and sheet_name not in a_wb.sheetnames:
        logger.warning(f"!AVISO! - No existe la hoja '{sheet_name}' en el excel '{file_xls}'. "
                       f"Se cogerá la primera hoja '{sh1}'")
        sheet_name = sh1

    if not sheet_name:
        sheet_name = sh1

    a_sheet = a_wb[sheet_name]

    return a_sheet


def header_sheet(a_sheet, col_ini=None, col_fin=None):
    """
    Retorna lista con los valores de la primera fila que se tomaran como cabecera

    Args:
        a_sheet (openpyxl.worksheet.woksheet):
        col_ini (int=None):
        col_fin (int=None):

    Returns:
        list
    """
    return [col if col else "" for col in next(a_sheet.iter_rows(1, 1, col_ini, col_fin, values_only=True), [])]


def iter_vals_sheet(a_sheet, row_header=False, row_ini=None, row_fin=None, col_ini=None, col_fin=None):
    """
    Itera las filas de la hoja indicada y devuelve las filas como diccionarios indexados por el nombre de columna.
    Si se indica row_header=True entonces se tomaran los valores de la primera fila como nombres de columna

    Args:
        a_sheet (openpyxl.worksheet.woksheet):
        row_header (bool=False):
        row_ini (int=None):
        row_fin (int=None):
        col_ini (int=None):
        col_fin (int=None):

    Yields:
        dict
    """
    nom_cols_head = None
    if row_header:
        nom_cols_head = header_sheet(a_sheet, col_ini, col_fin)
        if not row_ini or row_ini == 1:
            row_ini = 2

    for row in a_sheet.iter_rows(row_ini, row_fin, col_ini, col_fin, values_only=row_header):
        if row_header:
            yield dict(zip(nom_cols_head, row))
        else:
            yield {cell.column_letter: cell.value for cell in row}


def close_file_sheet(a_sheet):
    """
    Cierra la conexion al fichero excel del sheet pasado por parametro

    Args:
        a_sheet:
    """
    a_sheet.parent._archive.close()


if __name__ == '__main__':
    from fire import Fire
    Fire()