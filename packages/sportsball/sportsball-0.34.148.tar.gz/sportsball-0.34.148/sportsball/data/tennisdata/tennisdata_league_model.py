"""TennisData league model."""

import urllib.parse
from io import BytesIO
from typing import Any, Iterator
from urllib.parse import urlparse

import tqdm
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from scrapesession.scrapesession import ScrapeSession  # type: ignore

from ..game_model import GameModel
from ..league import League
from ..league_model import SHUTDOWN_FLAG, LeagueModel, needs_shutdown
from .tennisdata_game_model import create_tennisdata_game_model


class TennisDataLeagueModel(LeagueModel):
    """TennisData implementation of the league model."""

    def __init__(
        self,
        league: League,
        session: ScrapeSession,
        position: int | None = None,
    ) -> None:
        super().__init__(league, session, position=position)

    @classmethod
    def name(cls) -> str:
        """The name of the league model."""
        return "tennisdata-league-model"

    @classmethod
    def position_validator(cls) -> dict[str, str]:
        """Tennis position validators."""
        return {}

    def _row_to_game(self, row: Any) -> GameModel | None:
        tour_number_cell = str(row[0].value)
        if tour_number_cell in {"ATP", "None", "WTA"}:
            return None
        location_cell = str(row[1].value).strip()
        date_cell = str(row[3].value).strip()
        court_cell = str(row[5].value).strip()
        surface_cell = str(row[6].value).strip()
        best_of_cell = str(row[8].value).strip()
        if best_of_cell == "None":
            return None
        winner_cell = str(row[9].value).strip()
        loser_cell = str(row[10].value).strip()
        winner_rank_cell = (
            str(row[11].value).strip()
            if row[11].value is not None and row[11].value != "N/A"
            else None
        )
        loser_rank_cell = (
            str(row[12].value).strip()
            if row[12].value is not None and row[12].value != "N/A"
            else None
        )
        winner_total_points_cell = (
            str(row[13].value).strip()
            if row[13].value is not None and row[13].value != "N/A"
            else None
        )
        loser_total_points_cell = (
            str(row[14].value).strip()
            if row[14].value is not None and row[14].value != "N/A"
            else None
        )
        winner_points_set_1_cell = (
            str(row[15].value).strip() if row[15].value is not None else None
        )
        loser_points_set_1_cell = (
            str(row[16].value).strip() if row[16].value is not None else None
        )
        winner_points_set_2_cell = (
            str(row[17].value).strip() if row[17].value is not None else None
        )
        loser_points_set_2_cell = (
            str(row[18].value).strip() if row[18].value is not None else None
        )
        winner_points_set_3_cell = (
            str(row[19].value).strip() if row[19].value is not None else None
        )
        loser_points_set_3_cell = (
            str(row[20].value).strip() if row[20].value is not None else None
        )
        current_cell = 21
        winner_points_set_4_cell = None
        loser_points_set_4_cell = None
        winner_points_set_5_cell = None
        loser_points_set_5_cell = None
        if self.league == League.ATP:
            winner_points_set_4_cell = (
                str(row[current_cell].value).strip()
                if row[current_cell].value is not None
                else None
            )
            current_cell += 1
            loser_points_set_4_cell = (
                str(row[current_cell].value).strip()
                if row[current_cell].value is not None
                else None
            )
            current_cell += 1
            winner_points_set_5_cell = (
                str(row[current_cell].value).strip()
                if row[current_cell].value is not None
                else None
            )
            current_cell += 1
            loser_points_set_5_cell = (
                str(row[current_cell].value).strip()
                if row[current_cell].value is not None
                else None
            )
            current_cell += 1
        winner_sets_cell = (
            str(row[current_cell].value).strip()
            if row[current_cell].value is not None
            else None
        )
        current_cell += 1
        loser_sets_cell = (
            str(row[current_cell].value).strip()
            if row[current_cell].value is not None
            else None
        )
        current_cell += 2

        winner_odds_cell = None
        if row[current_cell].value is not None:
            winner_odds_cell = str(row[current_cell].value).strip()
            current_cell += 1
        else:
            current_cell += 6
            winner_odds_cell = str(row[current_cell].value).strip()
            current_cell += 1

        loser_odds_cell: str | None = str(row[current_cell].value).strip()
        if winner_odds_cell == "None":
            winner_odds_cell = None
        if loser_odds_cell == "None":
            loser_odds_cell = None

        current_cell += 1
        return create_tennisdata_game_model(
            location=location_cell,
            date=date_cell,
            court=court_cell,
            surface=surface_cell,
            best_of=best_of_cell,
            winner=winner_cell,
            loser=loser_cell,
            winner_rank=winner_rank_cell,
            loser_rank=loser_rank_cell,
            winner_total_points=winner_total_points_cell,
            loser_total_points=loser_total_points_cell,
            winner_points_set_1=winner_points_set_1_cell,
            loser_points_set_1=loser_points_set_1_cell,
            winner_points_set_2=winner_points_set_2_cell,
            loser_points_set_2=loser_points_set_2_cell,
            winner_points_set_3=winner_points_set_3_cell,
            loser_points_set_3=loser_points_set_3_cell,
            winner_points_set_4=winner_points_set_4_cell,
            loser_points_set_4=loser_points_set_4_cell,
            winner_points_set_5=winner_points_set_5_cell,
            loser_points_set_5=loser_points_set_5_cell,
            winner_sets=winner_sets_cell,
            loser_sets=loser_sets_cell,
            winner_odds=winner_odds_cell,
            loser_odds=loser_odds_cell,
            session=self.session,
            league=self.league,
        )

    @property
    def games(self) -> Iterator[GameModel]:
        """Find all the games."""
        with self.session.wayback_disabled():
            try:
                with tqdm.tqdm(position=self.position) as pbar:
                    url = "http://www.tennis-data.co.uk/alldata.php"
                    response = None
                    with self.session.cache_disabled():
                        response = self.session.get(url)
                    response.raise_for_status()
                    soup = BeautifulSoup(response.text, "lxml")
                    spreadsheet_urls = []
                    for a in soup.find_all("a", href=True):
                        if needs_shutdown():
                            return
                        spreadsheet_url = urllib.parse.urljoin(url, a.get("href"))
                        if not spreadsheet_url.endswith(".xlsx"):
                            continue
                        o = urlparse(spreadsheet_url)
                        path = o.path.split("/")
                        if self.league == League.WTA and path[-2].endswith("w"):
                            spreadsheet_urls.append(spreadsheet_url)
                        elif self.league == League.ATP and not path[-2].endswith("w"):
                            spreadsheet_urls.append(spreadsheet_url)

                    for count, spreadsheet_url in enumerate(
                        sorted(spreadsheet_urls, reverse=True)
                    ):
                        response = None
                        if count == 0:
                            with self.session.cache_disabled():
                                self.session.cache.delete(urls=[spreadsheet_url])
                                response = self.session.get(spreadsheet_url)
                        else:
                            response = self.session.get(spreadsheet_url)
                        response.raise_for_status()
                        workbook = load_workbook(filename=BytesIO(response.content))
                        ws = workbook.active
                        if ws is None:
                            raise ValueError("ws is null.")
                        for row in ws.iter_rows():
                            game_model = self._row_to_game(row)
                            if game_model is not None:
                                pbar.update(1)
                                pbar.set_description(f"TennisData - {game_model.dt}")
                                yield game_model
            except Exception as exc:
                SHUTDOWN_FLAG.set()
                raise exc
