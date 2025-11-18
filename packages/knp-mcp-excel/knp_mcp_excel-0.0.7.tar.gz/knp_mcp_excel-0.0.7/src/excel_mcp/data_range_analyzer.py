from typing import Any, Optional, Dict, List, Tuple
import logging
import os
import tempfile
import requests
from urllib.parse import urlparse
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .exceptions import ValidationError

logger = logging.getLogger(__name__)


def _is_url(path: str) -> bool:
    """ตรวจสอบว่า path เป็น URL หรือไม่"""
    try:
        result = urlparse(path)
        return all([result.scheme, result.netloc])
    except:
        return False


def _download_file(url: str) -> str:
    """
    ดาวน์โหลดไฟล์จาก URL และเก็บไว้ใน temp file
    
    Args:
        url: URL ของไฟล์
        
    Returns:
        path ของไฟล์ชั่วคราว
    """
    try:
        logger.info(f"Downloading file from: {url}")
        response = requests.get(url, timeout=30, stream=True)
        response.raise_for_status()
        
        # สร้าง temp file
        suffix = os.path.splitext(urlparse(url).path)[1] or '.xlsx'
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
        
        # เขียนข้อมูลลงไฟล์
        for chunk in response.iter_content(chunk_size=8192):
            if chunk:
                temp_file.write(chunk)
        
        temp_file.close()
        logger.info(f"File downloaded to: {temp_file.name}")
        return temp_file.name
        
    except requests.exceptions.RequestException as e:
        logger.error(f"Failed to download file from URL: {e}")
        raise ValidationError(f"Failed to download file from URL: {str(e)}")
    except Exception as e:
        logger.error(f"Error downloading file: {e}")
        raise ValidationError(f"Error downloading file: {str(e)}")


def _get_filepath(filepath_or_url: str) -> Tuple[str, bool]:
    """
    รับ filepath หรือ URL และคืนค่า local filepath พร้อมบอกว่าเป็น temp file หรือไม่
    
    Args:
        filepath_or_url: path หรือ URL ของไฟล์
        
    Returns:
        Tuple ของ (filepath, is_temp)
    """
    if _is_url(filepath_or_url):
        temp_path = _download_file(filepath_or_url)
        return temp_path, True
    return filepath_or_url, False


def get_used_range(filepath: str, sheet_name: str) -> Dict[str, Any]:
    """
    หาค่า data range ที่ใช้งานจริงใน sheet
    
    Args:
        filepath: path ของไฟล์ Excel หรือ URL
        sheet_name: ชื่อ sheet ที่ต้องการตรวจสอบ
        
    Returns:
        Dictionary ที่มีข้อมูล:
        - range: data range ในรูปแบบ "A1:Z100"
        - start_cell: cell เริ่มต้น
        - end_cell: cell สุดท้าย
        - rows: จำนวนแถว
        - columns: จำนวนคอลัมน์
        - total_cells: จำนวน cell ทั้งหมด
    """
    local_path, is_temp = _get_filepath(filepath)
    
    try:
        wb = load_workbook(local_path, data_only=True)
        
        if sheet_name not in wb.sheetnames:
            raise ValidationError(f"Sheet '{sheet_name}' not found in workbook")
        
        ws = wb[sheet_name]
        
        # หา min/max row และ column ที่มีข้อมูล
        min_row = ws.min_row
        max_row = ws.max_row
        min_col = ws.min_column
        max_col = ws.max_column
        
        # สร้าง cell reference
        start_cell = f"{get_column_letter(min_col)}{min_row}"
        end_cell = f"{get_column_letter(max_col)}{max_row}"
        data_range = f"{start_cell}:{end_cell}"
        
        # คำนวณจำนวน
        total_rows = max_row - min_row + 1
        total_cols = max_col - min_col + 1
        total_cells = total_rows * total_cols
        
        wb.close()
        
        return {
            "sheet": sheet_name,
            "range": data_range,
            "start_cell": start_cell,
            "end_cell": end_cell,
            "rows": total_rows,
            "columns": total_cols,
            "total_cells": total_cells,
            "dimensions": {
                "min_row": min_row,
                "max_row": max_row,
                "min_column": min_col,
                "max_column": max_col
            }
        }
        
    except ValidationError:
        raise
    except Exception as e:
        logger.error(f"Error analyzing data range: {e}")
        raise ValidationError(f"Failed to analyze data range: {str(e)}")
    finally:
        # ลบ temp file ถ้ามี
        if is_temp and os.path.exists(local_path):
            try:
                os.unlink(local_path)
            except:
                pass


def get_all_sheets_ranges(filepath: str) -> Dict[str, Any]:
    """
    หาค่า data range ของทุก sheet ในไฟล์
    
    Args:
        filepath: path ของไฟล์ Excel หรือ URL
        
    Returns:
        Dictionary ที่มีข้อมูลของทุก sheet
    """
    local_path, is_temp = _get_filepath(filepath)
    
    try:
        wb = load_workbook(local_path, data_only=True)
        
        results = {
            "file": filepath,
            "total_sheets": len(wb.sheetnames),
            "sheets": []
        }
        
        for sheet_name in wb.sheetnames:
            try:
                ws = wb[sheet_name]
                
                min_row = ws.min_row
                max_row = ws.max_row
                min_col = ws.min_column
                max_col = ws.max_column
                
                start_cell = f"{get_column_letter(min_col)}{min_row}"
                end_cell = f"{get_column_letter(max_col)}{max_row}"
                data_range = f"{start_cell}:{end_cell}"
                
                total_rows = max_row - min_row + 1
                total_cols = max_col - min_col + 1
                total_cells = total_rows * total_cols
                
                results["sheets"].append({
                    "sheet": sheet_name,
                    "range": data_range,
                    "start_cell": start_cell,
                    "end_cell": end_cell,
                    "rows": total_rows,
                    "columns": total_cols,
                    "total_cells": total_cells,
                    "dimensions": {
                        "min_row": min_row,
                        "max_row": max_row,
                        "min_column": min_col,
                        "max_column": max_col
                    }
                })
            except Exception as e:
                logger.warning(f"Could not analyze sheet '{sheet_name}': {e}")
                results["sheets"].append({
                    "sheet": sheet_name,
                    "error": str(e)
                })
        
        wb.close()
        return results
        
    except Exception as e:
        logger.error(f"Error analyzing workbook: {e}")
        raise ValidationError(f"Failed to analyze workbook: {str(e)}")
    finally:
        # ลบ temp file ถ้ามี
        if is_temp and os.path.exists(local_path):
            try:
                os.unlink(local_path)
            except:
                pass


def find_data_tables(
    filepath: str, 
    sheet_name: str,
    min_rows: int = 2,
    min_cols: int = 2
) -> List[Dict[str, Any]]:
    """
    ค้นหาตารางข้อมูลใน sheet (กลุ่มของ cell ที่มีข้อมูลต่อเนื่องกัน)
    
    Args:
        filepath: path ของไฟล์ Excel หรือ URL
        sheet_name: ชื่อ sheet
        min_rows: จำนวนแถวขั้นต่ำของตาราง
        min_cols: จำนวนคอลัมน์ขั้นต่ำของตาราง
        
    Returns:
        List ของตารางข้อมูลที่พบ
    """
    local_path, is_temp = _get_filepath(filepath)
    
    try:
        wb = load_workbook(local_path, data_only=True)
        
        if sheet_name not in wb.sheetnames:
            raise ValidationError(f"Sheet '{sheet_name}' not found")
        
        ws = wb[sheet_name]
        tables = []
        
        # สแกนหาบล็อกข้อมูลที่ต่อเนื่องกัน
        current_table = None
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            row_has_data = any(cell.value is not None for cell in row)
            
            if row_has_data:
                if current_table is None:
                    # เริ่มตารางใหม่
                    current_table = {
                        "start_row": row[0].row,
                        "end_row": row[0].row,
                        "start_col": None,
                        "end_col": None
                    }
                else:
                    # ขยายตารางปัจจุบัน
                    current_table["end_row"] = row[0].row
                
                # หา column range
                for cell in row:
                    if cell.value is not None:
                        if current_table["start_col"] is None:
                            current_table["start_col"] = cell.column
                            current_table["end_col"] = cell.column
                        else:
                            current_table["start_col"] = min(current_table["start_col"], cell.column)
                            current_table["end_col"] = max(current_table["end_col"], cell.column)
            else:
                # แถวว่าง - จบตาราง
                if current_table is not None:
                    # ตรวจสอบขนาดตาราง
                    rows = current_table["end_row"] - current_table["start_row"] + 1
                    cols = current_table["end_col"] - current_table["start_col"] + 1
                    
                    if rows >= min_rows and cols >= min_cols:
                        start_cell = f"{get_column_letter(current_table['start_col'])}{current_table['start_row']}"
                        end_cell = f"{get_column_letter(current_table['end_col'])}{current_table['end_row']}"
                        
                        tables.append({
                            "range": f"{start_cell}:{end_cell}",
                            "start_cell": start_cell,
                            "end_cell": end_cell,
                            "rows": rows,
                            "columns": cols,
                            "total_cells": rows * cols
                        })
                    
                    current_table = None
        
        # ตรวจสอบตารางสุดท้าย
        if current_table is not None:
            rows = current_table["end_row"] - current_table["start_row"] + 1
            cols = current_table["end_col"] - current_table["start_col"] + 1
            
            if rows >= min_rows and cols >= min_cols:
                start_cell = f"{get_column_letter(current_table['start_col'])}{current_table['start_row']}"
                end_cell = f"{get_column_letter(current_table['end_col'])}{current_table['end_row']}"
                
                tables.append({
                    "range": f"{start_cell}:{end_cell}",
                    "start_cell": start_cell,
                    "end_cell": end_cell,
                    "rows": rows,
                    "columns": cols,
                    "total_cells": rows * cols
                })
        
        wb.close()
        
        return {
            "sheet": sheet_name,
            "tables_found": len(tables),
            "tables": tables
        }
        
    except ValidationError:
        raise
    except Exception as e:
        logger.error(f"Error finding data tables: {e}")
        raise ValidationError(f"Failed to find data tables: {str(e)}")
    finally:
        # ลบ temp file ถ้ามี
        if is_temp and os.path.exists(local_path):
            try:
                os.unlink(local_path)
            except:
                pass


def get_column_data_info(
    filepath: str,
    sheet_name: str,
    data_range: Optional[str] = None
) -> Dict[str, Any]:
    """
    วิเคราะห์ข้อมูลในแต่ละคอลัมน์
    
    Args:
        filepath: path ของไฟล์ Excel หรือ URL
        sheet_name: ชื่อ sheet
        data_range: range ที่ต้องการวิเคราะห์ (ถ้าไม่ระบุจะใช้ทั้ง sheet)
        
    Returns:
        Dictionary ที่มีข้อมูลแต่ละคอลัมน์
    """
    local_path, is_temp = _get_filepath(filepath)
    
    try:
        wb = load_workbook(local_path, data_only=True)
        
        if sheet_name not in wb.sheetnames:
            raise ValidationError(f"Sheet '{sheet_name}' not found")
        
        ws = wb[sheet_name]
        
        # ถ้าไม่ระบุ range ให้ใช้ทั้งหมด
        if data_range:
            start_cell, end_cell = data_range.split(":")
            from .cell_utils import parse_cell_range
            start_row, start_col, end_row, end_col = parse_cell_range(start_cell, end_cell)
        else:
            start_row = ws.min_row
            end_row = ws.max_row
            start_col = ws.min_column
            end_col = ws.max_column
        
        columns_info = []
        
        for col in range(start_col, end_col + 1):
            col_letter = get_column_letter(col)
            
            # นับข้อมูลในคอลัมน์
            values = []
            non_empty_count = 0
            
            for row in range(start_row, end_row + 1):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value is not None:
                    non_empty_count += 1
                    values.append(cell_value)
            
            # ตรวจสอบประเภทข้อมูล
            data_types = set()
            for val in values:
                data_types.add(type(val).__name__)
            
            columns_info.append({
                "column": col_letter,
                "column_index": col,
                "header": ws.cell(row=start_row, column=col).value,
                "total_cells": end_row - start_row + 1,
                "non_empty_cells": non_empty_count,
                "empty_cells": (end_row - start_row + 1) - non_empty_count,
                "data_types": list(data_types)
            })
        
        wb.close()
        
        return {
            "sheet": sheet_name,
            "range": f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}",
            "total_columns": len(columns_info),
            "columns": columns_info
        }
        
    except ValidationError:
        raise
    except Exception as e:
        logger.error(f"Error analyzing column data: {e}")
        raise ValidationError(f"Failed to analyze column data: {str(e)}")
    finally:
        # ลบ temp file ถ้ามี
        if is_temp and os.path.exists(local_path):
            try:
                os.unlink(local_path)
            except:
                pass