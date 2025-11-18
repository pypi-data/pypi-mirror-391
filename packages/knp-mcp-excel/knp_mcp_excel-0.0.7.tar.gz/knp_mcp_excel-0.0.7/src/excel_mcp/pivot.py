from typing import Any
import uuid
import logging

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font

from .data import read_excel_range
from .cell_utils import parse_cell_range
from .exceptions import ValidationError, PivotError

logger = logging.getLogger(__name__)

def create_pivot_table(
    filepath: str,
    sheet_name: str,
    data_range: str,
    rows: list[str],
    values: list[str],
    columns: list[str] | None = None,
    agg_func: str = "sum",
    pivot_name: str = "PivotTable",
    subtotal_rows: str = "false",
    filter_null: str = "true"
) -> dict[str, Any]:
    """Create pivot table in sheet using Excel table functionality
    
    Args:
        filepath: Path to Excel file
        sheet_name: Name of worksheet containing source data
        data_range: Source data range reference
        rows: Fields for row labels
        values: Fields for values
        columns: Optional fields for column labels
        agg_func: Aggregation function (sum, count, count_all, average, max, min)
        pivot_name: Name for the pivot sheet
        subtotal_rows: "true" = show subtotal rows for each group, "false" = hide subtotals (default: "false")
        filter_null: Whether to filter out rows with all zero values - "true" or "false" (default: "true")
        
    Returns:
        Dictionary with status message and pivot table dimensions
    """
    try:
        # Convert string parameters to boolean
        subtotal_rows_bool = str(subtotal_rows).lower() in ("true", "1", "yes")
        filter_null_bool = str(filter_null).lower() in ("true", "1", "yes")
        wb = load_workbook(filepath)
        if sheet_name not in wb.sheetnames:
            raise ValidationError(f"Sheet '{sheet_name}' not found")
        
        # Parse ranges
        if ':' not in data_range:
            raise ValidationError("Data range must be in format 'A1:B2'")
            
        try:
            start_cell, end_cell = data_range.split(':')
            start_row, start_col, end_row, end_col = parse_cell_range(start_cell, end_cell)
        except ValueError as e:
            raise ValidationError(f"Invalid data range format: {str(e)}")
            
        if end_row is None or end_col is None:
            raise ValidationError("Invalid data range format: missing end coordinates")
            
        # Create range string
        data_range_str = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
        
        # Clean up field names by removing aggregation suffixes
        def clean_field_name(field: str) -> str:
            field = str(field).strip()
            for suffix in [" (sum)", " (average)", " (count)", " (min)", " (max)"]:
                if field.lower().endswith(suffix):
                    return field[:-len(suffix)]
            return field

        # Read source data and convert to list of dicts
        try:
            data_as_list = read_excel_range(filepath, sheet_name, start_cell, end_cell)
            if not data_as_list or len(data_as_list) < 2:
                raise PivotError("Source data must have a header row and at least one data row.")
            
            headers = [str(h) for h in data_as_list[0]]
            data = [dict(zip(headers, row)) for row in data_as_list[1:]]

            if not data:
                raise PivotError("No data rows found after header.")

        except Exception as e:
            raise PivotError(f"Failed to read or process source data: {str(e)}")

        # Validate aggregation function
        valid_agg_funcs = ["sum", "average", "count", "count_all", "min", "max"]
        if agg_func.lower() not in valid_agg_funcs:
            raise ValidationError(
                f"Invalid aggregation function. Must be one of: {', '.join(valid_agg_funcs)}"
            )

        # Validate field names exist in data
        if data:
            available_fields_raw = data[0].keys()
            available_fields = {clean_field_name(str(header)).lower() for header in available_fields_raw}
            
            for field_list, field_type in [(rows, "row"), (values, "value")]:
                for field in field_list:
                    if clean_field_name(str(field)).lower() not in available_fields:
                        raise ValidationError(
                            f"Invalid {field_type} field '{field}'. "
                            f"Available fields: {', '.join(sorted(available_fields_raw))}"
                        )

            if columns:
                for field in columns:
                    if clean_field_name(str(field)).lower() not in available_fields:
                        raise ValidationError(
                            f"Invalid column field '{field}'. "
                            f"Available fields: {', '.join(sorted(available_fields_raw))}"
                        )

        # Clean up field names
        cleaned_rows = [clean_field_name(field) for field in rows]
        cleaned_values = [clean_field_name(field) for field in values]
        cleaned_columns = [clean_field_name(field) for field in columns] if columns else None

        # Create pivot sheet
        pivot_sheet_name = f"{pivot_name}"
        if pivot_sheet_name in wb.sheetnames:
            wb.remove(wb[pivot_sheet_name])
        pivot_ws = wb.create_sheet(pivot_sheet_name)

        # Get unique values for row and column fields
        row_field_values = {}
        for field in cleaned_rows:
            all_values = []
            for record in data:
                value = record.get(field, '')
                if value is None:
                    value = ''
                all_values.append(str(value).strip())
            row_field_values[field] = sorted(set(filter(None, all_values)))

        col_field_values = {}
        if cleaned_columns:
            for field in cleaned_columns:
                all_values = []
                for record in data:
                    value = record.get(field, '')
                    if value is None:
                        value = ''
                    all_values.append(str(value).strip())
                col_field_values[field] = sorted(set(filter(None, all_values)))

        # Generate row and column combinations
        row_combinations = _get_combinations(row_field_values)
        col_combinations = _get_combinations(col_field_values) if col_field_values else [{}]

        # Build header row
        current_row = 1
        current_col = 1
        
        # Write row field headers
        for field in cleaned_rows:
            cell = pivot_ws.cell(row=current_row, column=current_col, value=field)
            cell.font = Font(bold=True)
            current_col += 1
        
        # Write column headers
        if cleaned_columns:
            for col_combo in col_combinations:
                for value_field in cleaned_values:
                    col_label_parts = [str(col_combo.get(f, '')) for f in cleaned_columns]
                    col_label = ' - '.join(col_label_parts) if col_label_parts else ''
                    header_text = f"{col_label}" if col_label else f"{value_field}"
                    cell = pivot_ws.cell(row=current_row, column=current_col, value=header_text)
                    cell.font = Font(bold=True)
                    current_col += 1
            
            # Add Grand Total column(s)
            for value_field in cleaned_values:
                cell = pivot_ws.cell(row=current_row, column=current_col, value="Grand Total")
                cell.font = Font(bold=True)
                current_col += 1
        else:
            for value_field in cleaned_values:
                cell = pivot_ws.cell(row=current_row, column=current_col, value=f"{value_field} ({agg_func})")
                cell.font = Font(bold=True)
                current_col += 1

        # Group row combinations by first row field for subtotals
        grouped_rows = {}
        if len(cleaned_rows) > 0:
            first_field = cleaned_rows[0]
            for row_combo in row_combinations:
                key = row_combo[first_field]
                if key not in grouped_rows:
                    grouped_rows[key] = []
                grouped_rows[key].append(row_combo)
        else:
            grouped_rows['_all'] = row_combinations
        
        # Write data rows
        current_row = 2
        rows_written = 0
        
        for group_key in sorted(grouped_rows.keys()):
            group_rows = grouped_rows[group_key]
            group_row_count = 0
            
            # Initialize group subtotals
            if cleaned_columns:
                group_subtotals = {}
                for col_combo in col_combinations:
                    for value_field in cleaned_values:
                        key = (tuple(col_combo.items()), value_field)
                        group_subtotals[key] = 0
                group_grand_totals = {vf: 0 for vf in cleaned_values}
            else:
                group_subtotals = {vf: 0 for vf in cleaned_values}
            
            # Process each row in the group
            for row_combo in group_rows:
                if cleaned_columns:
                    # Calculate values for each column combination
                    row_totals = {value_field: 0 for value_field in cleaned_values}
                    row_values = []
                    
                    for col_combo in col_combinations:
                        filtered_data = _filter_data(data, row_combo, col_combo)
                        
                        for value_field in cleaned_values:
                            try:
                                value = _aggregate_values(filtered_data, value_field, agg_func)
                                row_values.append(value)
                                row_totals[value_field] += value
                                key = (tuple(col_combo.items()), value_field)
                                group_subtotals[key] += value
                            except Exception as e:
                                raise PivotError(f"Failed to aggregate values for field '{value_field}': {str(e)}")
                    
                    for value_field in cleaned_values:
                        group_grand_totals[value_field] += row_totals[value_field]
                    
                    # Check if should filter this row (based on filter_null parameter)
                    should_filter = False
                    if filter_null_bool:
                        all_zeros = all(v == 0 for v in row_values) and all(row_totals[vf] == 0 for vf in cleaned_values)
                        should_filter = all_zeros
                    
                    if not should_filter:
                        col = 1
                        for field in cleaned_rows:
                            pivot_ws.cell(row=current_row, column=col, value=row_combo[field])
                            col += 1
                        
                        for value in row_values:
                            pivot_ws.cell(row=current_row, column=col, value=value)
                            col += 1
                        
                        for value_field in cleaned_values:
                            pivot_ws.cell(row=current_row, column=col, value=row_totals[value_field])
                            col += 1
                        
                        current_row += 1
                        rows_written += 1
                        group_row_count += 1
                else:
                    # No columns - simple aggregation
                    filtered_data = _filter_data(data, row_combo, {})
                    
                    row_values = []
                    for value_field in cleaned_values:
                        try:
                            value = _aggregate_values(filtered_data, value_field, agg_func)
                            row_values.append(value)
                            group_subtotals[value_field] += value
                        except Exception as e:
                            raise PivotError(f"Failed to aggregate values for field '{value_field}': {str(e)}")
                    
                    # Check if should filter this row (based on filter_null parameter)
                    should_filter = False
                    if filter_null_bool:
                        all_zeros = all(v == 0 for v in row_values)
                        should_filter = all_zeros
                    
                    if not should_filter:
                        col = 1
                        for field in cleaned_rows:
                            pivot_ws.cell(row=current_row, column=col, value=row_combo[field])
                            col += 1
                        
                        for value in row_values:
                            pivot_ws.cell(row=current_row, column=col, value=value)
                            col += 1
                        
                        current_row += 1
                        rows_written += 1
                        group_row_count += 1
            
            # Write subtotal row for this group (only if subtotal_rows is true)
            if subtotal_rows_bool and group_row_count > 0 and len(cleaned_rows) > 0:
                col = 1
                cell = pivot_ws.cell(row=current_row, column=col, value=f"{group_key} Total")
                cell.font = Font(bold=True)
                col += 1
                
                for _ in range(len(cleaned_rows) - 1):
                    col += 1
                
                if cleaned_columns:
                    for col_combo in col_combinations:
                        for value_field in cleaned_values:
                            key = (tuple(col_combo.items()), value_field)
                            cell = pivot_ws.cell(row=current_row, column=col, value=group_subtotals[key])
                            cell.font = Font(bold=True)
                            col += 1
                    
                    for value_field in cleaned_values:
                        cell = pivot_ws.cell(row=current_row, column=col, value=group_grand_totals[value_field])
                        cell.font = Font(bold=True)
                        col += 1
                else:
                    for value_field in cleaned_values:
                        cell = pivot_ws.cell(row=current_row, column=col, value=group_subtotals[value_field])
                        cell.font = Font(bold=True)
                        col += 1
                
                current_row += 1
                rows_written += 1
        
        # Write Grand Total row (always shown)
        if rows_written > 0:
            col = 1
            cell = pivot_ws.cell(row=current_row, column=col, value="Grand Total")
            cell.font = Font(bold=True)
            col += 1
            
            for _ in range(len(cleaned_rows) - 1):
                col += 1
            
            if cleaned_columns:
                for col_combo in col_combinations:
                    for value_field in cleaned_values:
                        filtered_data = _filter_data(data, {}, col_combo)
                        value = _aggregate_values(filtered_data, value_field, agg_func)
                        cell = pivot_ws.cell(row=current_row, column=col, value=value)
                        cell.font = Font(bold=True)
                        col += 1
                
                for value_field in cleaned_values:
                    filtered_data = _filter_data(data, {}, {})
                    value = _aggregate_values(filtered_data, value_field, agg_func)
                    cell = pivot_ws.cell(row=current_row, column=col, value=value)
                    cell.font = Font(bold=True)
                    col += 1
            else:
                for value_field in cleaned_values:
                    filtered_data = _filter_data(data, {}, {})
                    value = _aggregate_values(filtered_data, value_field, agg_func)
                    cell = pivot_ws.cell(row=current_row, column=col, value=value)
                    cell.font = Font(bold=True)
                    col += 1
            
            rows_written += 1

        # Calculate table dimensions
        total_rows = rows_written
        if cleaned_columns:
            total_cols = len(cleaned_rows) + (len(col_combinations) * len(cleaned_values)) + len(cleaned_values)
        else:
            total_cols = len(cleaned_rows) + len(cleaned_values)

        # Create a table for the pivot data
        try:
            pivot_range = f"A1:{get_column_letter(total_cols)}{total_rows}"
            pivot_table = Table(
                displayName=f"PivotTable_{uuid.uuid4().hex[:8]}", 
                ref=pivot_range
            )
            style = TableStyleInfo(
                name="None",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=True
            )
            pivot_table.tableStyleInfo = style
            pivot_ws.add_table(pivot_table)
        except Exception as e:
            raise PivotError(f"Failed to create pivot table formatting: {str(e)}")

        try:
            wb.save(filepath)
        except Exception as e:
            raise PivotError(f"Failed to save workbook: {str(e)}")
        
        return {
            "message": "Summary table created successfully",
            "details": {
                "source_range": data_range_str,
                "pivot_sheet": pivot_sheet_name,
                "rows": cleaned_rows,
                "columns": cleaned_columns or [],
                "values": cleaned_values,
                "aggregation": agg_func,
                "subtotal_rows_shown": subtotal_rows_bool,
                "null_rows_filtered": filter_null_bool
            }
        }
        
    except (ValidationError, PivotError) as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to create pivot table: {e}")
        raise PivotError(str(e))


def _get_combinations(field_values: dict[str, list]) -> list[dict]:
    """Get all combinations of field values."""
    if not field_values:
        return [{}]
    
    result = [{}]
    for field, values in list(field_values.items()):
        new_result = []
        for combo in result:
            for value in sorted(values):
                new_combo = combo.copy()
                new_combo[field] = value
                new_result.append(new_combo)
        result = new_result
    return result


def _filter_data(data: list[dict], row_filters: dict, col_filters: dict) -> list[dict]:
    """Filter data based on row and column filters."""
    result = []
    for record in data:
        matches = True
        for field, value in row_filters.items():
            record_value = record.get(field, '')
            if record_value is None:
                record_value = ''
            if str(record_value).strip() != str(value).strip():
                matches = False
                break
        for field, value in col_filters.items():
            record_value = record.get(field, '')
            if record_value is None:
                record_value = ''
            if str(record_value).strip() != str(value).strip():
                matches = False
                break
        if matches:
            result.append(record)
    return result


def _aggregate_values(data: list[dict], field: str, agg_func: str) -> float:
    """Aggregate values using the specified function."""
    if agg_func == "count":
        return len([record for record in data if field in record and record[field] is not None])
    
    if agg_func == "count_all":
        return len([record for record in data if field in record])
    
    # Filter numeric values only
    values = [record[field] for record in data if field in record and isinstance(record[field], (int, float))]
    if not values:
        return 0
        
    if agg_func == "sum":
        return sum(values)
    elif agg_func == "average":
        return sum(values) / len(values)
    elif agg_func == "min":
        return min(values)
    elif agg_func == "max":
        return max(values)
    else:
        return sum(values)  # Default to sum