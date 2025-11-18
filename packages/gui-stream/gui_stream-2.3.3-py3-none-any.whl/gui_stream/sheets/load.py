#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from __future__ import annotations
import sys
from abc import ABC, abstractmethod
from typing import Optional, List
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet._read_only import ReadOnlyWorksheet
from pandas.io.parsers import TextFileReader
import pandas as pd
from gui_stream.app_ui.core.progress import ProgressBarAdapter, ProgressBarSimple
from soup_files import File, Directory


class ABCSheetReader(ABC):
    """Classe abstrata para leitura de planilhas CSV e EXCEL"""
    def __init__(self, file_path: File, *, progress_bar: ProgressBarAdapter = None):
        if progress_bar is None:
            self.progress_bar: ProgressBarAdapter = ProgressBarAdapter(ProgressBarSimple())
        else:
            self.progress_bar: ProgressBarAdapter = progress_bar
        self.file_path:File = file_path
        self.df = None
        self._running:bool = False
        self.sheet_name: str = None   

    @abstractmethod
    def read(self):
        pass

    @abstractmethod
    def get_sheet_names(self) -> list[str]:
        pass
    
    def is_running(self) -> bool:
        return self._running

    def get_dataframe(self) -> pd.DataFrame:
        return self.df


class CSVReader(ABCSheetReader):
    def __init__(self, path:File, *, progress: ProgressBarAdapter = None, separator:str = '\t'):
        super().__init__(path, progress_bar=progress)
        self.separator:str=separator

    def get_sheet_names(self):
        return [self.file_path.name()]
    
    def read(self):
        self._running = True
        self.progress_bar.start()
        self.progress_bar.update(0, "Iniciando leitura do CSV")

        total_lines = sum(1 for _ in open(self.file_path.absolute(), encoding="utf-8")) - 1  # excluir header
        chunks:List[TextFileReader] = []
        for i, chunk in enumerate(pd.read_csv(self.file_path.absolute(), chunksize=1000, sep=self.separator)):
            chunks.append(chunk)
            percent: float = ((i + 1) * 1000) / total_lines * 100
            self.progress_bar.update(percent, f"Lendo CSV [{self.file_path.basename()}]")

        self.df = pd.concat(chunks, ignore_index=True)
        self.progress_bar.update(100, "Leitura finalizada!", )
        self.progress_bar.stop()
        self._running = False


class ExcelReader(ABCSheetReader):
    def __init__(self, path, *, sheet_name:str=None, progress: ProgressBarAdapter = None):
        super().__init__(path, progress_bar=progress)
        self._num_rows: int = 0
        self._num_columns: int = 0
        self.__wb: Workbook = None
        self.__read_only_wsheet: ReadOnlyWorksheet = None
        self.__sheet_names: list[str] = []
        self.__origin_header: List[str] = []
        self.__sheet_name:str = sheet_name
        self.initValues()

    @property
    def sheet_name(self) -> str | None:
        return self.__sheet_name
    
    @sheet_name.setter
    def sheet_name(self, new: str):
        self.__sheet_name = new
        self.initValues()

    def initValues(self):
        # Carregar o Workbook
        self.progress_bar.start()
        self.progress_bar.update_text('Iniciando leitura, aguarde!!!')
        try:
            self.__wb: Workbook = load_workbook(self.file_path.absolute(), read_only=True)
        except Exception as e:
            print(f'{__class__.__name__}\n{e}\n')

        if self.__wb is None:
            return
        # WorkSheet
        if self.sheet_name is None:
            self.__read_only_wsheet = self.__wb.active
        else:
            self.__read_only_wsheet = self.__wb[self.sheet_name]
        self.__sheet_names = self.__wb.sheetnames

        # Número de linhas, colunas e cabeçalho
        if self.__read_only_wsheet is None:
            return
        self._num_rows = self.__read_only_wsheet.max_row
        self._num_columns = self.__read_only_wsheet.max_column
        self.__origin_header = next(self.__read_only_wsheet.iter_rows(values_only=True))
                
    def get_header(self) -> List[str]:
        """Cabeçalho original da planilha"""
        return self.__origin_header
    
    def get_num_rows(self) -> int:
        """Retorna o número de linhas de uma planilha excel."""
        return self._num_rows
    
    def get_num_columns(self) -> int:
        return self._num_columns
    
    def get_current_workbook(self) -> Workbook | None:
        """Retorna um objeto Workbook da planilha ativa."""
        return self.__wb
        
    def get_read_only_work_sheet(self) -> ReadOnlyWorksheet | None:
        """Objeto para obter algumas propriedades da planilha."""
        return self.__read_only_wsheet
    
    def get_sheet_names(self):
        return self.__sheet_names
    
    def read(self):
        self._running = True
        self.progress_bar.start()
        self.progress_bar.update(0, "Iniciando leitura do Excel")
        
        list_data:List[tuple] = []
        if self.get_read_only_work_sheet() is None:
            self.progress_bar.update(0, "Falha na leitura")
            self._running = False
            return
        #
        _rows: tuple = self.get_read_only_work_sheet().iter_rows(values_only=True)
        maxnum: int = self.get_num_rows()
        for num, row in enumerate(_rows, 1):
            self.progress_bar.update(
                    ((num+1)/maxnum) * 100,
                    f'Lendo planilha: {self.get_read_only_work_sheet().title} | Linhas [{num+1} de {maxnum}]'
                )
            list_data.append(row)
            #sys.stdout.flush()
            
        if len(list_data) < 1:
            self.df = pd.DataFrame()
        else:
            self.df = pd.DataFrame(list_data)
    
        if (self.get_header() is not None) and (self.get_header() != []):
            self.df.columns = self.get_header()
            
        self.progress_bar.update(100, 'Operação finalizada!!!')
        print()
        self.progress_bar.stop()
        self._running = False
     


class SheetInputStream(object):
    def __init__(
                    self, 
                    path: File, 
                    *,
                    sheet_name: str = None, 
                    progress: ProgressBarAdapter = ProgressBarAdapter(ProgressBarSimple()), 
                    separator='\t',
                ):
        
        self.path:File = path
        self.sheet_name:str = sheet_name
        self.separator: str = separator
        self.progress: ProgressBarAdapter = progress
        
    def get_reader(self) -> ABCSheetReader:
        if self.path.is_csv():
            return CSVReader(self.path, progress=self.progress, separator=self.separator)
        elif self.path.is_excel():
            return ExcelReader(self.path, sheet_name=self.sheet_name, progress=self.progress)
        else:
            print('-------------------------------------------------------')
            print(self.path.extension())
            raise ValueError(f"Formato de arquivo não suportado: \n{self.path.absolute()}")

    def get_sheet_names(self) -> list[str]:
        return self.get_reader().get_sheet_names()

    def read(self, sheet_name=None) -> pd.DataFrame:
        self.progress.start()
        if sheet_name is not None:
            self.sheet_name = sheet_name
        reader: ABCSheetReader = self.get_reader()
        reader.read()
        self.progress.stop()
        return reader.get_dataframe()
    
    def read_sheets(self) -> dict[str, pd.DataFrame]:
        data_sheets: dict[str, pd.DataFrame] = {}
        names = self.get_sheet_names()
        
        for name in names:
            data_sheets[name] = self.read(name)
        return data_sheets



