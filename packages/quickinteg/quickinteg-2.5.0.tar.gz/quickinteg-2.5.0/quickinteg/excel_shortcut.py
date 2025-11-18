#! python3  # noqa: E265

################################################################################
# This file is part of quickinteg.

# quickinteg is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# quickinteg is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with quickinteg.  If not, see <https://www.gnu.org/licenses/>.
################################################################################

# ############################################################################
# ########## Libraries #############
# ##################################

# standard library
import logging

# 3rd party - try embedded first (for QGIS), fallback to normal import (for CLI)
try:
    from .external import openpyxl
except ImportError:
    import openpyxl

# package
from . import processing_logging

# ############################################################################
# ########## Globals ###############
# ##################################

module_logger = logging.getLogger(__name__)
module_logger.setLevel(logging.DEBUG)


# ############################################################################
# ########## Functions #############
# ##################################


def log(message: str, logLevel: str = "info"):
    """Map log message on different logging pipes. Mainly: QGIS Processing log system.

    Args:
        message (str): [description]
        logLevel (str, optional): [description]. Defaults to "info".
    """
    processing_logging.log(message, module_logger, logLevel)


def read_sheet_to_dict(
    excel_file_path, list_necessary_columns, primary_key=None, excel_sheet_name=None
):
    """
    Lecture d'un fichier excel vert un dict de dict.
    Chaque ligne est un dict avec comme clés les en-têtes de colones.
    Le fichier est un dict de lignes avec comme clé la valeur \
    dans la colonne "primary_key".
    Si primary_key == None, on utilise le numéro de ligne
    ex :

    le tableau

    Primary_key     |   colonne1    |   colonne2
    key1            |   value1      |   value2
    key2            |   value3      |   value4

    donnera en sortie un dict
    {
        key1 {
            Primary_key : key1,
            colonne1 : value1,
            colonne2 : value2
            },
        key2 {
            Primary_key : key2,
            colonne1 : value3,
            colonne2 : value4
            }
        }
    """
    wb = openpyxl.load_workbook(excel_file_path, data_only=True)
    sheet = wb[excel_sheet_name] if excel_sheet_name else wb.active

    rows = sheet.iter_rows(values_only=True)
    headers = next(rows)  # Get the first row with column names
    header_set = set(headers)

    necessary_columns_set = set(list_necessary_columns)
    if not necessary_columns_set.issubset(header_set):
        missing_columns = necessary_columns_set - header_set
        raise ValueError(
            f"The necessary columns {missing_columns} were not found in the Excel sheet"
        )

    header_index = {header: headers.index(header) for header in list_necessary_columns}

    dict_output = {}
    for i, row in enumerate(rows, start=1):  # start=1 because Excel is 1-indexed
        row_data = {header: row[header_index[header]] for header in list_necessary_columns}
        key = row_data.get(primary_key) if primary_key else i
        dict_output[key] = row_data

    return dict_output


def read_all_sheet_to_dict(excel_file_path, list_necessary_columns, primary_key):
    dict_output = dict()
    wb = openpyxl.load_workbook(excel_file_path, data_only=True)
    for sheet_name in wb.sheetnames:
        dict_output.update(
            read_sheet_to_dict(
                excel_file_path,
                list_necessary_columns,
                primary_key,
                excel_sheet_name=sheet_name,
            )
        )
    return dict_output


def read_sheet_to_tab(excel_file_path, excel_sheet_name=None):
    """Lecture d'un fichier excel et retour un dict de deux listes :
    une liste header
    une double liste data"""
    # Ouverture du xlsx
    wb = openpyxl.load_workbook(excel_file_path, data_only=True)

    # Si pas de sheet_name passé en paramètre, on ouvre le premier onglet
    if excel_sheet_name is None:
        excel_sheet = wb[wb.sheetnames[0]]
    # Sinon on ouvre l'onglet passé en paramètre
    else:
        excel_sheet = wb[excel_sheet_name]

    # Stockage des valeurs dans un tableau
    header = []
    data = []
    for i, row in enumerate(excel_sheet.iter_rows(values_only=True)):
        if i == 0:
            header.extend(row)
        else:
            data.append(list(row))
    return {"header": header, "data": data}


# ############################################################################
# ########## Main ##################
# ##################################

# for compatibility after simplifying name
read_excel_sheet_to_dict = read_sheet_to_dict

# for compatibility after simplifying name
read_excel_sheet_to_tab = read_sheet_to_tab
